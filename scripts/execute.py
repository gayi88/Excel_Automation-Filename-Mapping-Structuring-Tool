import re
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# =========================================================
# PATHS
# =========================================================
SOURCE_FILE = Path(r"D:\Excel_mapping_project\input\master.xlsm")
OUTPUT_FILE = Path(r"D:\Excel_mapping_project\input\master_output.xlsm")

TARGET_SHEET = "Handlingsförteckning"

# =========================================================
# CONFIG FOR BOTH SHEETS
# =========================================================

SHEETS = [

    # -------- STATION --------
    {
        "NAME": "Station",
        "COL_FILENAME": 9,
        "COL_TYPE": 10,
        "TEKNIK_MAP": {
            "C": "Teknikövergripande",
            "N": "Miljö",
            "Q": "Kvalitet",
            "A": "Arkitekt",
            "B": "Brand",
            "E": "El",
            "H": "Hiss",
            "J": "Akustik",
            "K": "Konstruktion",
            "M": "Mark",
            "S": "Styr och övervakning",
            "T": "Tele (fastighet)",
            "V": "VS och Sprinkler",
            "X": "Berg (fastighet)",
        },
        "CODE_RE": re.compile(r"^[A-ZÅÄÖ]+")
    },

    # -------- BEST --------
    {
        "NAME": "BEST & Anläggning",
        "COL_FILENAME": 7,
        "COL_TYPE": 8,
        "TEKNIK_MAP": {
            "C": "Teknikövergripande",
            "N": "Miljö",
            "Q": "Kvalitet",
            "b": "Bana/Spår",
            "e": "El",
            "f": "Fjärrstyrning",
            "k": "Kanalisation",
            "s": "Signal",
            "t": "Tele",
            "u": "Kabel",
            "v": "El 230V och 400V",
            "d": "Konstbyggnad",
            "g": "Geoteknik",
            "j": "Akustik ute (buller)",
            "l": "Landskap",
            "m": "Mark",
            "p": "Grundkarta/Baskarta",
            "r": "VA",
            "si": "Skalskydd inbrottslarm",
            "w": "Ledningssamordning",
            "x": "Berg",
            "z": "Mätning/geodesi",
        },
        "CODE_RE": re.compile(r"^[A-Za-zÅÄÖåäö]+")
    }
]

SECTION_RE = re.compile(r"^\d+\s")
SUB_UP_RE  = re.compile(r"^\d+\.[A-ZÅÄÖ]")
SUB_LO_RE  = re.compile(r"^\d+\.[a-zåäö]")

# =========================================================
def make_copy():
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(f"Source file not found: {SOURCE_FILE}")
    shutil.copy(SOURCE_FILE, OUTPUT_FILE)

# =========================================================
def extract_section_number(text):
    try:
        return int(str(text).split()[0])
    except:
        return None

# =========================================================
def find_section_row(ws, section_number):
    prefix = f"{section_number} "
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 2).value
        if val and str(val).strip().startswith(prefix):
            return r
    return None

# =========================================================
def find_subsection_row(ws, section_number, teknik_code):

    if section_number == 0:
        target = f"{teknik_code} "
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, 2).value
            if val and str(val).strip().startswith(target):
                return r
        return None

    target = f"{section_number}.{teknik_code}"

    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 2).value
        if val and str(val).strip().startswith(target):
            return r

    return None

# =========================================================
def create_subsection(ws, section_row, section_number, teknik_code, teknik_map):

    insert_row = section_row + 1

    while insert_row <= ws.max_row:
        val = ws.cell(insert_row, 2).value
        if val and SECTION_RE.match(str(val)):
            break
        insert_row += 1

    ws.insert_rows(insert_row)

    name = teknik_map.get(teknik_code, teknik_code)

    ws.cell(insert_row, 2).value = f"{teknik_code} {name}"

    return insert_row

# =========================================================
def file_exists(ws, subsection_row, filename):

    r = subsection_row + 1
    safety_limit = ws.max_row + 200

    while r <= safety_limit:
        val = ws.cell(r, 2).value
        if val and (SECTION_RE.match(str(val)) or SUB_UP_RE.match(str(val)) or SUB_LO_RE.match(str(val))):
            break

        if ws.cell(r, 1).value == filename:
            return True
        r += 1

    return False

# =========================================================
def find_write_row(ws, subsection_row):

    r = subsection_row + 1
    safety_limit = ws.max_row + 200

    while r <= safety_limit:

        val = ws.cell(r, 2).value
        if val and (SECTION_RE.match(str(val)) or SUB_UP_RE.match(str(val)) or SUB_LO_RE.match(str(val))):
            ws.insert_rows(r)
            return r

        if not ws.cell(r, 1).value and not ws.cell(r, 2).value:
            return r

        r += 1

    ws.insert_rows(ws.max_row + 1)
    return ws.max_row

# =========================================================
def style_subsection_row(ws, row):
    fill = PatternFill(fill_type="solid", start_color="FCE4D6")
    font = Font(bold=True)
    align = Alignment(horizontal="left")

    for col in range(1, 7):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align

# =========================================================
def process_sheet(wb, config):

    if config["NAME"] not in wb.sheetnames:
        print(f"⚠ Sheet not found: {config['NAME']}")
        return

    master_ws = wb[config["NAME"]]
    target_ws = wb[TARGET_SHEET]

    for r in range(2, master_ws.max_row + 1):

        filename = master_ws.cell(r, config["COL_FILENAME"]).value
        doc_type = master_ws.cell(r, config["COL_TYPE"]).value

        if not filename or not doc_type:
            continue

        filename = str(filename).strip()
        section_number = extract_section_number(doc_type)
        if section_number is None:
            continue

        match = config["CODE_RE"].match(filename)
        if not match:
            continue

        teknik_code = match.group(0)
        if teknik_code not in config["TEKNIK_MAP"]:
            continue

        section_row = find_section_row(target_ws, section_number)
        if not section_row:
            continue

        subsection_row = find_subsection_row(target_ws, section_number, teknik_code)

        if not subsection_row:
            subsection_row = create_subsection(
                target_ws, section_row, section_number, teknik_code, config["TEKNIK_MAP"]
            )

        # 🎨 STYLE ALWAYS (new + existing)
        style_subsection_row(target_ws, subsection_row)

        if file_exists(target_ws, subsection_row, filename):
            continue

        write_row = find_write_row(target_ws, subsection_row)

        target_ws.row_dimensions[write_row].height = 16

        target_ws.cell(write_row, 1).value = filename

        print(f"Inserted → {filename}")

# =========================================================
def main():

    make_copy()

    wb = load_workbook(OUTPUT_FILE, keep_vba=True, data_only=True)

    for sheet_config in SHEETS:
        print(f"\nProcessing {sheet_config['NAME']}...")
        process_sheet(wb, sheet_config)

    wb.save(OUTPUT_FILE)

    print(f"\n✅ MERGED FILE CREATED → {OUTPUT_FILE}")

# =========================================================
if __name__ == "__main__":
    main()